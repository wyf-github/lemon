'''
    - -coding= utf-8- -
    @authour=Wangyafei
    @time=2019/3/12 9:45
    @file=do_excel.py
'''
import openpyxl
import pprint
from openpyxl import load_workbook
from project1.common import project_path
from project1.common.read_config import ReadConfig


class DoExcel:
  def __init__(self,file_name,sheet_name):
      self.file_name=file_name
      self.sheet_name=sheet_name

  def read_data(self,section):
      try:
          case_id=ReadConfig(project_path.conf_path).get_data(section,'case_id')
          wb=load_workbook(self.file_name)
          sheet=wb[self.sheet_name]
      except Exception as e:
         print("加载文件失败{}".format(e))

      tel = self.get_tel()
      test_data=[]

      for i in range(2,sheet.max_row+1):
            row_data={}
            row_data['CaseId']=sheet.cell(i,1).value
            row_data['Module'] = sheet.cell(i, 2).value
            row_data['Title'] = sheet.cell(i, 3).value
            row_data['Url'] = sheet.cell(i, 4).value
            row_data['Method'] = sheet.cell(i, 5).value
            if sheet.cell(i, 6).value.find('tel')!=-1: #找到tel的子字符串

                row_data['Params'] = sheet.cell(i, 6).value.replace('tel', str(tel))#replace是字符串直接的替换，字符串和整数之间不能替换
                self.update_tel(int(tel+1))
            else: #如果没有tel字符串就不用替换
                row_data['Params'] = sheet.cell(i, 6).value

            row_data['ExpectedResult'] = sheet.cell(i, 7).value
            test_data.append(row_data)
      wb.close()
      final_data=[]

      if case_id =='all': #执行所有用例
          final_data=test_data
      else:
          for i in case_id:  #遍历case_id里的值
              final_data.append(test_data[i-1])
      return final_data

  def get_tel(self):
      wb = load_workbook(self.file_name)
      sheet = wb['tel']
      wb.close()
      return int(sheet.cell(1, 2).value)


  def update_tel(self,new_tel):
      '''写回手机号码'''
      wb = load_workbook(self.file_name)
      sheet = wb['tel']
      sheet.cell(1, 2,new_tel)
      wb.save(self.file_name)
      wb.close()

  def write_data(self,row,col,value):
        '''写数据到表中'''
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        sheet.cell(row,col).value=value
        wb.save(self.file_name)
        wb.close()



if __name__=='__main__':
    file_name=project_path.testcases_path
    sheet_name='1'
    # test_data=DoExcel(project_path.testcases_path,sheet_name).read_data('RechargeCASE')
    test_data2 = DoExcel(project_path.testcases_path,sheet_name ).write_data(2,2,'a')
    # pprint.pprint(test_data)